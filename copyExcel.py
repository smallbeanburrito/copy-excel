from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
import sys
import csv

i = 0  # index for iterating through copyColumns
rowStart = 5  # row number of field names
valueStart = 18  # row number where values begin on source workbook
addOrDelete = ''  # column determining if a row should be copied
copyColumns = [] # list of column names that should be copied

args = len(sys.argv)

addOrDelete = sys.argv[1]
selectColumns = sys.argv[2]
source = sys.argv[3]

if args == 5:
    destination = sys.argv[4]
else:
    destination = "destination.xlsx"

with open(selectColumns, newline='') as csvFile:
    copyColumns = list(csv.reader(csvFile))[0]

try:
    #  copy from this workbook
    src = load_workbook(filename=source, data_only=True)
    SS = src.active
    copyCols = SS.iter_cols(min_row=valueStart)

except FileNotFoundError:
    print("You entered: " + source)
    print("Enter the correct excel file IDIOT!")

if args == 5:
    # copy to this workbook
    des = load_workbook(filename=destination)
    DS = des.active
else:
    #create new workbook
    des = Workbook()
    DS = des.active


def valid_fields(field, name):  # checks if a column should be copied
    if field == name:
        return True
    else:
        return False


# holds a list of rows to be copied
addRow = []
addField = (SS[addOrDelete])
for cell in addField:
    if cell.value == 'add':
        addRow.append(cell.row)

# holds a list of columns to be copied
addColumn = []
for col in copyCols:
    if i < len(copyColumns):
        if valid_fields(copyColumns[i], SS[col[0].column_letter + str(rowStart)].value):
            i += 1
            addColumn.append(col[0].column_letter)

absCol = 1  # begin copying at this column of destination workbook
for column in addColumn:
    absRow = 2  # being copying at this row of destination workbook
    for row in addRow:
        if args != 5:
            # add column names
            DS[get_column_letter(absCol) + '1'] = copyColumns[absCol-1]
        DS[get_column_letter(absCol) + str(absRow)] = SS[column + str(row)].value
        absRow += 1
    absCol += 1

des.save(destination)
des.close()
src.close()
print("Excel sheet copied")

